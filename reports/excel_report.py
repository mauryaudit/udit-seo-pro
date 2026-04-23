"""
reports/excel_report.py
Generates a comprehensive Excel workbook with 7 sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ── Style helpers ──────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def _border():
    s = Side(style="thin", color="DDDDEE")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _score_fill(score):
    if score >= 80: return _fill("D4EDDA")
    elif score >= 60: return _fill("FFF3CD")
    return _fill("F8D7DA")

def _score_font(score):
    if score >= 80: return _font(bold=True, color="1A7A3A")
    elif score >= 60: return _font(bold=True, color="DD7700")
    return _font(bold=True, color="CC2200")

def _sev_fill(sev):
    if sev == "critical": return _fill("F8D7DA")
    if sev == "warning": return _fill("FFF3CD")
    return _fill("D1ECF1")

def _sev_font(sev):
    if sev == "critical": return _font(bold=True, color="CC2200", size=9)
    if sev == "warning": return _font(bold=True, color="DD7700", size=9)
    return _font(bold=True, color="2244AA", size=9)

def _header_row(ws, row, headers, widths, bg="1A1A2E"):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(bg)
        c.alignment = _center()
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 18


MODULE_LABELS = {
    "title": "Title Tag",
    "meta_description": "Meta Description",
    "headings": "Heading Structure",
    "content": "Content Quality (E-E-A-T)",
    "images": "Image Optimization",
    "schema": "Schema / Structured Data",
    "technical": "Technical SEO",
    "open_graph": "Open Graph / Social",
    "links": "Internal & External Links",
    "robots_txt": "Robots.txt",
    "sitemap": "XML Sitemap",
    "geo_aeo": "GEO / AI Search",
    "core_web_vitals": "Core Web Vitals Signals",
    "mobile": "Mobile SEO",
    "international": "International SEO",
    "security": "Security",
}

PHASE_FILLS = {
    1: "1A1A2E",
    2: "1A4A8A",
    3: "0A6A3A",
    4: "6A1A6A",
}


def generate_excel(results, phase_plan, semrush_data, output_path,
                   eeat_result=None, content_brief=None):
    wb = openpyxl.Workbook()

    _sheet_summary(wb, results)
    _sheet_issues(wb, results)
    _sheet_passes(wb, results)
    _sheet_semrush(wb, semrush_data)
    _sheet_phase_plan(wb, phase_plan)
    if eeat_result:
        _sheet_eeat(wb, eeat_result)
    if content_brief:
        _sheet_content_brief(wb, content_brief)
    _sheet_raw_data(wb, results)
    _sheet_chart(wb, results)

    wb.save(output_path)
    print(f"  Excel saved: {output_path}")


# ── SHEET 1: Executive Summary ─────────────────────────

def _sheet_summary(wb, results):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    # Banner
    for row in [1, 2, 3]:
        ws.merge_cells(f"A{row}:G{row}")
    ws["A1"].value = "SEO AUDIT REPORT"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = _fill("1A1A2E")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 36

    ws["A2"].value = results["url"]
    ws["A2"].font = _font(color="AAAACC", size=10)
    ws["A2"].fill = _fill("1A1A2E")
    ws["A2"].alignment = _center()

    ws["A3"].value = f"Audit Date: {results['audit_date']}"
    ws["A3"].font = _font(color="AAAACC", size=9)
    ws["A3"].fill = _fill("1A1A2E")
    ws["A3"].alignment = _center()
    ws.row_dimensions[3].height = 18

    # Score cards
    overall = results["overall_score"]
    critical_c = len([i for i in results["all_issues"] if i["severity"] == "critical"])
    warning_c = len([i for i in results["all_issues"] if i["severity"] == "warning"])
    info_c = len([i for i in results["all_issues"] if i["severity"] == "info"])
    pass_c = len(results["all_passes"])

    cards = [
        ("A5:B6", overall, "Overall Score", _score_fill(overall), "000000" if overall >= 60 else "CC2200"),
        ("C5:C6", critical_c, "Critical", _fill("F8D7DA"), "CC2200"),
        ("D5:D6", warning_c, "Warnings", _fill("FFF3CD"), "DD7700"),
        ("E5:E6", info_c, "Info", _fill("D1ECF1"), "2244AA"),
        ("F5:G6", pass_c, "Passed", _fill("D4EDDA"), "1A7A3A"),
    ]
    ws.row_dimensions[5].height = 50
    ws.row_dimensions[6].height = 50
    for rng, val, label, fill, fc in cards:
        ws.merge_cells(rng)
        start = rng.split(":")[0]
        c = ws[start]
        c.value = f"{val}\n{label}"
        c.font = Font(bold=True, size=14, name="Calibri", color=fc)
        c.fill = fill
        c.alignment = _center()
        c.border = _border()

    # Module scores table
    row = 8
    _header_row(ws, row, ["Module", "Score", "Issues", "Critical", "Warnings", "Status"], [32, 12, 12, 12, 12, 14])
    for mod_key, label in MODULE_LABELS.items():
        mod = results["modules"].get(mod_key, {})
        s = mod.get("score", 0)
        issues = mod.get("issues", [])
        crit = len([i for i in issues if i["severity"] == "critical"])
        warn = len([i for i in issues if i["severity"] == "warning"])
        status = "Good" if s >= 80 else ("Review" if s >= 60 else "Fix Now")
        row += 1
        for col, val in enumerate([label, s, len(issues), crit, warn, status], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _border()
            c.font = _font(size=9)
            c.alignment = _left() if col == 1 else _center()
            if col == 2:
                c.fill = _score_fill(s)
                c.font = _score_font(s)
            elif row % 2 == 0:
                c.fill = _fill("F5F5FC")


# ── SHEET 2: All Issues ────────────────────────────────

def _sheet_issues(wb, results):
    ws = wb.create_sheet("All Issues")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value = "All Issues & Recommendations"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = _fill("1A1A2E")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 24

    _header_row(ws, 2, ["#", "Severity", "Module", "Issue / Finding", "Action Required"], [6, 14, 26, 70, 24])

    row = 3
    for sev in ["critical", "warning", "info"]:
        for issue in results["all_issues"]:
            if issue["severity"] != sev:
                continue
            action = {"critical": "Fix immediately", "warning": "Fix within 1 week", "info": "Improve when possible"}[sev]
            vals = [row - 2, sev.upper(), MODULE_LABELS.get(issue["module"], issue["module"]), issue["msg"], action]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = _border()
                c.font = _font(size=9)
                c.alignment = _left()
                if col == 2:
                    c.fill = _sev_fill(sev)
                    c.font = _sev_font(sev)
                    c.alignment = _center()
                elif col == 1:
                    c.alignment = _center()
                elif row % 2 == 0:
                    c.fill = _fill("FAFAFA")
            row += 1


# ── SHEET 3: Passed Checks ────────────────────────────

def _sheet_passes(wb, results):
    ws = wb.create_sheet("Passed Checks")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Passed Checks"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = _fill("1A7A3A")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 24

    _header_row(ws, 2, ["#", "Module", "Check Passed"], [6, 26, 80], bg="1A7A3A")

    for i, p in enumerate(results["all_passes"], 1):
        row = i + 2
        for col, val in enumerate([i, MODULE_LABELS.get(p["module"], p["module"]), p["msg"]], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _border()
            c.font = _font(size=9, color="1A7A3A")
            c.alignment = _left()
            if i % 2 == 0:
                c.fill = _fill("E8F5E9")
        ws.cell(row=row, column=1).alignment = _center()


# ── SHEET 4: Semrush Data ─────────────────────────────

def _sheet_semrush(wb, semrush_data):
    ws = wb.create_sheet("Semrush Data")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Semrush Data"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = _fill("FF6900")  # Semrush brand color
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 24

    row = 2
    if not any(v.get("ok") for v in semrush_data.values() if isinstance(v, dict)):
        ws.cell(row=row, column=1, value="Semrush API key not configured. Add SEMRUSH_API_KEY to .env file.")
        ws.cell(row=row, column=1).font = _font(color="CC2200", size=10, italic=True)
        ws.column_dimensions["A"].width = 70
        return

    for section_name, data in semrush_data.items():
        if not isinstance(data, dict) or not data.get("ok") or not data.get("data"):
            continue

        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=section_name.replace("_", " ").title())
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill("333355")
        c.alignment = _left()
        row += 1

        # Write headers from first row keys
        rows_data = data["data"]
        if rows_data:
            headers = list(rows_data[0].keys())[:6]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = _font(bold=True, color="FFFFFF", size=8)
                c.fill = _fill("1A4A8A")
                c.border = _border()
                ws.column_dimensions[get_column_letter(col)].width = 22
            row += 1

            for data_row in rows_data[:20]:
                for col, key in enumerate(headers, 1):
                    c = ws.cell(row=row, column=col, value=str(data_row.get(key, ""))[:60])
                    c.font = _font(size=8)
                    c.border = _border()
                    if row % 2 == 0:
                        c.fill = _fill("F5F5FC")
                row += 1

        row += 1


# ── SHEET 5: Phase Plan ───────────────────────────────

def _sheet_phase_plan(wb, phase_plan):
    ws = wb.create_sheet("4-Phase SEO Plan")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"].value = "4-Phase SEO Action Plan"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = _fill("1A1A2E")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15

    row = 2
    for phase_num, phase_data in phase_plan.items():
        fill_hex = PHASE_FILLS[phase_num]

        # Phase header
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=f"Phase {phase_num}: {phase_data['name']}  |  {phase_data['timeframe']}  |  {phase_data['goal']}")
        c.font = _font(bold=True, color="FFFFFF", size=11)
        c.fill = _fill(fill_hex)
        c.alignment = _left()
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers
        for col, h in enumerate(["#", "Task", "Source / Data", "Effort", "Impact", "Status"], 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = _font(bold=True, color="FFFFFF", size=8)
            c.fill = _fill(fill_hex)
            c.border = _border()
            c.alignment = _center()
        ws.row_dimensions[row].height = 16
        row += 1

        for i, task in enumerate(phase_data.get("tasks", []), 1):
            vals = [i, task["task"], task.get("source", ""), task.get("effort", ""), task.get("impact", ""), ""]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(size=8)
                c.border = _border()
                c.alignment = _left()
                if i % 2 == 0:
                    c.fill = _fill("F5F5FC")
                if col == 1:
                    c.alignment = _center()
            row += 1

        row += 1


# ── SHEET 6: Raw Data ─────────────────────────────────

def _sheet_raw_data(wb, results):
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Raw Audit Data"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = _fill("1A1A2E")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    row = 2
    for mod_key, label in MODULE_LABELS.items():
        mod = results["modules"].get(mod_key, {})
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.fill = _fill("333355")
        c.alignment = _left()
        row += 1

        for k, v in mod.items():
            if k in {"issues", "passes"}:
                continue
            k_c = ws.cell(row=row, column=1, value=k)
            v_c = ws.cell(row=row, column=2, value=str(v)[:200] if isinstance(v, (list, dict)) else v)
            k_c.font = _font(size=8, bold=True)
            v_c.font = _font(size=8)
            k_c.fill = _fill("F5F5FC")
            k_c.border = _border()
            v_c.border = _border()
            row += 1
        row += 1


# ── SHEET 7: Score Chart ──────────────────────────────

def _sheet_chart(wb, results):
    ws = wb.create_sheet("Score Chart")
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value="Module").font = _font(bold=True)
    ws.cell(row=1, column=2, value="Score").font = _font(bold=True)

    chart_row = 2
    for mod_key, label in MODULE_LABELS.items():
        s = results["modules"].get(mod_key, {}).get("score", 0)
        ws.cell(row=chart_row, column=1, value=label)
        ws.cell(row=chart_row, column=2, value=s)
        chart_row += 1

    chart = BarChart()
    chart.type = "bar"
    chart.title = "SEO Module Scores"
    chart.style = 10
    chart.y_axis.title = "Score (/100)"
    chart.x_axis.title = "Module"
    chart.width = 24
    chart.height = 16

    data = Reference(ws, min_col=2, min_row=1, max_row=chart_row - 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=chart_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12


# ── SHEET 8: CORE-EEAT Scorecard ─────────────────────────

def _sheet_eeat(wb, eeat_result):
    if not eeat_result:
        return
    ws = wb.create_sheet("CORE-EEAT Scorecard")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value = f"CORE-EEAT Content Quality — Score: {eeat_result['total_score']}/100 — Grade {eeat_result['grade']}"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = _fill("1A1A2E")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    # Dimension summary
    row = 2
    dim_labels = {"C": "Comprehensiveness", "O": "Originality", "R": "Relevance", "E": "E-E-A-T"}
    for col, (d, dl) in enumerate(dim_labels.items(), 1):
        s = eeat_result["summary"].get(d, 0)
        c = ws.cell(row=row, column=col, value=f"{dl}\n{s}/100")
        c.font = _font(bold=True, size=11)
        c.fill = _score_fill(s)
        c.alignment = _center()
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[2].height = 36
    row = 3

    dims = eeat_result.get("dimensions", {})
    for d_key, d_label in dim_labels.items():
        dd = dims.get(d_key, {})
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"{d_label} — {dd.get('score',0)}/100")
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.fill = _fill("333355")
        row += 1
        _header_row(ws, row, ["ID", "Check", "Result", "Notes"], [10, 55, 14, 30], bg="1A4A8A")
        row += 1
        for item in dd.get("items", []):
            sv = item["score"]
            s_str = "Pass" if sv >= 1 else ("Partial" if sv >= 0.5 else "Fail")
            s_col = "1A7A3A" if sv >= 1 else ("DD7700" if sv >= 0.5 else "CC2200")
            vals = [item["id"], item["label"], s_str, item.get("note", "")]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(size=8, color=s_col if col == 3 else "000000", bold=(col==3))
                c.border = _border()
                c.alignment = _left()
                if row % 2 == 0:
                    c.fill = _fill("F5F5FC")
            row += 1
        row += 1


# ── SHEET 9: Content Brief ────────────────────────────────

def _sheet_content_brief(wb, content_brief):
    if not content_brief:
        return
    ws = wb.create_sheet("Content Brief")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    ws["A1"].value = "Content Brief & Optimization Guide"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = _fill("1A4A8A")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70

    cb = content_brief
    rows = [
        ("URL", cb.get("url", "")[:100]),
        ("Current H1", cb.get("current_h1", "")[:100]),
        ("Primary Keyword", cb.get("primary_keyword", "")),
        ("Secondary Keywords", ", ".join(cb.get("secondary_keywords", []))),
        ("LSI / Semantic Terms", ", ".join(cb.get("lsi_terms", []))),
        ("Current Word Count", str(cb.get("current_word_count", 0))),
        ("Target Word Count", cb.get("target_word_count", "")),
        ("CORE-EEAT Total", f"{cb.get('eeat_total','N/A')}/100 — Grade {cb.get('content_grade','N/A')}"),
        ("C / O / R / E Scores",
         f"C:{cb['eeat_scores'].get('C',0)}  O:{cb['eeat_scores'].get('O',0)}  R:{cb['eeat_scores'].get('R',0)}  E:{cb['eeat_scores'].get('E',0)}"),
    ]
    for i, (k, v) in enumerate(rows, 2):
        kc = ws.cell(row=i, column=1, value=k)
        vc = ws.cell(row=i, column=2, value=v)
        kc.font = _font(bold=True, size=9)
        vc.font = _font(size=9)
        kc.border = _border()
        vc.border = _border()
        kc.fill = _fill("F5F5FC") if i % 2 == 0 else _fill("FFFFFF")
        vc.fill = _fill("F5F5FC") if i % 2 == 0 else _fill("FFFFFF")

    row = len(rows) + 3

    def section(title, items, row):
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.fill = _fill("1A4A8A")
        row += 1
        for i, item in enumerate(items, 1):
            nc = ws.cell(row=row, column=1, value=str(i))
            ic = ws.cell(row=row, column=2, value=item)
            nc.font = _font(bold=True, size=8)
            ic.font = _font(size=8)
            nc.alignment = _center()
            ic.border = _border()
            nc.border = _border()
            if i % 2 == 0:
                ic.fill = _fill("F5F5FC")
            row += 1
        return row + 1

    row = section("Recommended H2 Structure", cb.get("recommended_h2_structure", []), row)
    row = section("Schema Markup to Add", cb.get("schema_to_implement", []), row)
    row = section("GEO / AI Search Optimizations", cb.get("geo_optimizations", []), row)
    row = section("Priority Actions (from Audit)", cb.get("priority_actions", []), row)
